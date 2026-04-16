import argparse
import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.run_human_benchmark import DEFAULT_SAMPLE_PATHS, MODEL_NAME, load_context
from marking_pipeline.core import (
    build_assessment_map,
    build_submission_texts_from_path,
    detect_submission_parts,
    extract_assessment_structure,
    refine_submission_granularity,
    segment_submission_parts,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export extracted assessment structure and downstream units for inspection."
    )
    parser.add_argument("--sample", default=str(DEFAULT_SAMPLE_PATHS[0]), help="Submission path to inspect.")
    parser.add_argument(
        "--output",
        default=str(Path("output") / "assessment_structure_inspection.xlsx"),
        help="Output Excel workbook path.",
    )
    return parser.parse_args()


def _flatten_structure(sections, depth: int = 0) -> list[tuple[int, object]]:
    rows: list[tuple[int, object]] = []
    for section in sections:
        rows.append((depth, section))
        if section.children:
            rows.extend(_flatten_structure(section.children, depth + 1))
    return rows


def main() -> None:
    args = parse_args()
    sample_path = Path(args.sample)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    context = load_context()
    structure_sections = extract_assessment_structure(context)
    assessment_map = build_assessment_map(context)

    script_text, structure_text = build_submission_texts_from_path(sample_path)
    structure_input_text = structure_text or script_text
    detected_parts = detect_submission_parts(
        structure_input_text,
        sample_path.name,
        MODEL_NAME,
        context=context,
    )
    segmented_parts = segment_submission_parts(script_text, detected_parts)
    refined_parts = refine_submission_granularity(segmented_parts, context, sample_path.name, MODEL_NAME)

    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    structure_sheet = workbook.create_sheet("ContextStructure")
    unit_sheet = workbook.create_sheet("AssessmentMap")
    parts_sheet = workbook.create_sheet("SubmissionParts")

    overview_sheet.append(["field", "value"])
    overview_rows = [
        ("sample_path", str(sample_path)),
        ("sample_file_name", sample_path.name),
        ("structure_sections_top_level", len(structure_sections)),
        ("structure_sections_total", len(_flatten_structure(structure_sections))),
        ("assessment_units", len(assessment_map.units)),
        ("detected_parts", len(detected_parts)),
        ("segmented_parts", len(segmented_parts)),
        ("refined_parts", len(refined_parts)),
        ("script_word_count", len(script_text.split())),
        ("structure_word_count", len(structure_input_text.split())),
    ]
    for row in overview_rows:
        overview_sheet.append(list(row))

    structure_sheet.append(
        [
            "depth",
            "label",
            "parent_label",
            "max_mark",
            "weight_text",
            "question_text_exact",
            "marking_instructions_exact",
            "anchor_phrases",
            "evidence_expectations",
            "dependency_labels",
        ]
    )
    for depth, section in _flatten_structure(structure_sections):
        structure_sheet.append(
            [
                depth,
                section.label,
                section.parent_label,
                section.max_mark,
                section.weight_text,
                section.question_text_exact,
                section.marking_instructions_exact,
                "\n".join(section.anchor_phrases),
                "\n".join(section.evidence_expectations),
                "\n".join(section.dependency_labels),
            ]
        )

    unit_sheet.append(
        [
            "label",
            "parent_label",
            "max_mark",
            "grading_mode",
            "task_type",
            "dependency_group",
            "question_text_exact",
            "marking_guidance",
            "rubric_text",
        ]
    )
    for unit in assessment_map.units:
        unit_sheet.append(
            [
                unit.label,
                unit.parent_label,
                unit.max_mark,
                unit.grading_mode,
                unit.task_type,
                unit.dependency_group,
                unit.question_text_exact,
                unit.marking_guidance,
                unit.rubric_text,
            ]
        )

    parts_sheet.append(
        [
            "stage",
            "label",
            "max_mark",
            "task_type",
            "question_text_exact",
            "marking_guidance",
            "section_word_count",
            "section_text",
        ]
    )
    for stage_name, part_list in (
        ("detected", detected_parts),
        ("segmented", segmented_parts),
        ("refined", refined_parts),
    ):
        for part in part_list:
            parts_sheet.append(
                [
                    stage_name,
                    part.label,
                    part.max_mark,
                    part.task_type,
                    part.question_text_exact,
                    part.marking_guidance,
                    len(part.section_text.split()),
                    part.section_text,
                ]
            )

    for sheet in (overview_sheet, structure_sheet, unit_sheet, parts_sheet):
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            letter = column[0].column_letter
            if letter in {"A", "B", "C", "D", "E"}:
                continue
            sheet.column_dimensions[letter].width = 32

    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
