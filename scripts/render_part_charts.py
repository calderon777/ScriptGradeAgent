import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_SUITE_DIR = ROOT_DIR / "output" / "end_to_end_suite_full_ec3040_2026-04-19_v2"
DEFAULT_OUTPUT = DEFAULT_SUITE_DIR / "human_vs_model_by_part.html"
DEFAULT_WORKBOOK = Path(
    r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040 deanonymised grades.xlsx"
)
PART_PATTERN = re.compile(r"(Part\s+\d+(?:\s+Q\d+)?):\s*([0-9]+(?:\.[0-9]+)?)/[0-9]+(?:\.[0-9]+)?", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render sorted human-vs-model charts for each question part.")
    parser.add_argument("--suite-dir", default=str(DEFAULT_SUITE_DIR), help="Suite output directory.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Workbook path with human marks.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output HTML path.")
    return parser.parse_args()


def parse_part_breakdown(text: str) -> dict[str, float]:
    breakdown: dict[str, float] = {}
    for label, score in PART_PATTERN.findall(text or ""):
        cleaned = re.sub(r"\s+", " ", label.strip()).title().replace("Q", "Q")
        breakdown[cleaned] = float(score)
    return breakdown


def load_human_parts(workbook_path: Path) -> dict[str, dict[str, float]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Regular submission point"]
    records: dict[str, dict[str, float]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        identifier, _full_name, _status, grade, feedback_comments = row[:5]
        if not identifier or grade is None:
            continue
        match = re.search(r"(\d{7})", str(identifier))
        if not match:
            continue
        records[match.group(1)] = parse_part_breakdown(str(feedback_comments or ""))
    return records


def slugify(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-") or "part"


def scale_x(index: int, total: int, left: float, width: float) -> float:
    if total <= 1:
        return left + width / 2
    return left + (index * width / (total - 1))


def scale_y(value: float, min_value: float, max_value: float, top: float, height: float) -> float:
    if max_value <= min_value:
        return top + height / 2
    ratio = (value - min_value) / (max_value - min_value)
    return top + height - (ratio * height)


def build_chart_svg(rows: list[dict[str, object]], title: str) -> str:
    width = 1100
    height = 420
    margin_left = 64
    margin_right = 20
    margin_top = 36
    margin_bottom = 54
    plot_width = width - margin_left - margin_right
    plot_height = height - margin_top - margin_bottom
    human_values = [float(row["human"]) for row in rows]
    model_values = [float(row["model"]) for row in rows]
    min_value = min(human_values + model_values + [0.0])
    max_value = max(human_values + model_values) + 2.0
    tick_step = 2
    tick_max = int(((max_value + tick_step - 1) // tick_step) * tick_step)
    ticks = list(range(0, max(tick_max, tick_step) + 1, tick_step))

    human_points: list[str] = []
    model_dots: list[str] = []
    grid: list[str] = []

    for tick in ticks:
        y = scale_y(float(tick), min_value, max_value, margin_top, plot_height)
        grid.append(
            f"<line x1=\"{margin_left}\" y1=\"{y:.2f}\" x2=\"{width - margin_right}\" y2=\"{y:.2f}\" stroke=\"#d1d5db\" stroke-width=\"1\" />"
        )
        grid.append(
            f"<text x=\"{margin_left - 10}\" y=\"{y + 4:.2f}\" text-anchor=\"end\" fill=\"#374151\" font-size=\"12\">{tick:g}</text>"
        )

    for index, row in enumerate(rows):
        x = scale_x(index, len(rows), margin_left, plot_width)
        y_human = scale_y(float(row["human"]), min_value, max_value, margin_top, plot_height)
        y_model = scale_y(float(row["model"]), min_value, max_value, margin_top, plot_height)
        human_points.append(f"{x:.2f},{y_human:.2f}")
        tooltip = (
            f"{row['participant_id']} | {Path(str(row['path'])).name} | "
            f"human={float(row['human']):.1f} | model={float(row['model']):.1f}"
        )
        model_dots.append(
            f"<circle cx=\"{x:.2f}\" cy=\"{y_model:.2f}\" r=\"4\" fill=\"#2563eb\"><title>{tooltip}</title></circle>"
        )

    return f"""
    <section class="chart-section" id="{slugify(title)}">
      <h2>{title}</h2>
      <p>{len(rows)} submissions with parsable human part marks.</p>
      <svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-label="{title}">
        <rect x="{margin_left}" y="{margin_top}" width="{plot_width}" height="{plot_height}" fill="#ffffff" stroke="#9ca3af" stroke-width="1" />
        {''.join(grid)}
        <line x1="{margin_left}" y1="{margin_top + plot_height}" x2="{width - margin_right}" y2="{margin_top + plot_height}" stroke="#111827" stroke-width="1.2" />
        <line x1="{margin_left}" y1="{margin_top}" x2="{margin_left}" y2="{margin_top + plot_height}" stroke="#111827" stroke-width="1.2" />
        <polyline points="{' '.join(human_points)}" fill="none" stroke="#dc2626" stroke-width="2.5" />
        {''.join(model_dots)}
        <text x="{width / 2:.2f}" y="{height - 16}" text-anchor="middle" fill="#374151" font-size="13">Submission rank sorted by human mark for {title}</text>
      </svg>
    </section>
    """


def main() -> None:
    args = parse_args()
    suite_dir = Path(args.suite_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    human_parts = load_human_parts(Path(args.workbook))
    run_dir = suite_dir / "cached_map_only" / "run_1"

    per_label: dict[str, list[dict[str, object]]] = {}
    for debug_path in sorted(run_dir.glob("*/debug_run.json")):
        payload = json.loads(debug_path.read_text(encoding="utf-8"))
        participant_id = str(payload["participant_id"])
        human_breakdown = human_parts.get(participant_id, {})
        part1_model_total = 0.0
        part1_found = False
        for row in payload.get("part_debug_rows", []):
            label = str(row["label"])
            if label in {"Part 1 Q1", "Part 1 Q2", "Part 1 Q3"}:
                part1_model_total += float(row["score"] or 0.0)
                part1_found = True
            if label not in human_breakdown:
                continue
            per_label.setdefault(label, []).append(
                {
                    "participant_id": participant_id,
                    "path": str(payload["path"]),
                    "human": float(human_breakdown[label]),
                    "model": float(row["score"] or 0.0),
                }
            )
        if part1_found and "Part 1" in human_breakdown:
            per_label.setdefault("Part 1", []).append(
                {
                    "participant_id": participant_id,
                    "path": str(payload["path"]),
                    "human": float(human_breakdown["Part 1"]),
                    "model": round(part1_model_total, 2),
                }
            )

    ordered_labels = sorted(
        per_label.keys(),
        key=lambda value: (int(re.search(r"Part\s+(\d+)", value).group(1)), int(re.search(r"Q(\d+)", value).group(1)) if "Q" in value else 0)
    )
    charts: list[str] = []
    toc: list[str] = []
    for label in ordered_labels:
        rows = sorted(per_label[label], key=lambda item: (float(item["human"]), str(item["participant_id"])))
        toc.append(f'<li><a href="#{slugify(label)}">{label}</a> ({len(rows)})</li>')
        charts.append(build_chart_svg(rows, label))

    html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Human vs Model by Part</title>
  <style>
    body {{
      font-family: Arial, sans-serif;
      margin: 24px;
      color: #111827;
      background: #ffffff;
    }}
    h1, p {{
      margin: 0 0 12px 0;
    }}
    ul {{
      columns: 3;
      padding-left: 20px;
      margin: 8px 0 24px 0;
    }}
    li {{
      margin: 4px 0;
    }}
    .legend {{
      display: flex;
      gap: 20px;
      margin: 8px 0 16px 0;
      font-size: 14px;
    }}
    .legend span {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
    }}
    .swatch {{
      width: 14px;
      height: 14px;
      display: inline-block;
    }}
    .chart-section {{
      margin: 28px 0 40px 0;
      border-top: 1px solid #e5e7eb;
      padding-top: 20px;
    }}
  </style>
</head>
<body>
  <h1>Sorted Human Marks vs Model Marks by Question Part</h1>
  <p>For each part, the red line is the human mark sorted low to high. Blue dots are model marks for the same scripts.</p>
  <div class="legend">
    <span><span class="swatch" style="background:#dc2626;"></span>Human marks</span>
    <span><span class="swatch" style="background:#2563eb; border-radius:50%;"></span>Model marks</span>
  </div>
  <ul>
    {''.join(toc)}
  </ul>
  {''.join(charts)}
</body>
</html>
"""
    output_path.write_text(html_text, encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
