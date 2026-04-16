import argparse
import json
import re
import sys
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from marking_pipeline.core import (  # noqa: E402
    DEFAULT_OLLAMA_URL,
    MarkingContext,
    _run_part_analysis_with_retry,
    apply_prepared_artifact_to_submission_parts,
    apply_assessment_map_to_submission_parts,
    build_missing_part_analysis,
    build_local_final_result,
    build_submission_diagnostics,
    build_submission_texts_from_path,
    clear_prepared_assessment_map_cache,
    detect_submission_parts,
    describe_moderation_plan,
    describe_structure_detection_mode,
    moderate_linked_part_analyses,
    prepare_assessment_map,
    read_path_text,
    read_paths_text,
    refine_submission_granularity,
    segment_submission_parts,
)


ROOT_DIR = Path(__file__).resolve().parents[1]
LAST_INGEST_DIR = ROOT_DIR / ".scriptgrade_cache" / "last_ingest"
LAST_INGEST_MANIFEST = LAST_INGEST_DIR / "manifest.json"
DEFAULT_WORKBOOK = Path(
    r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040 deanonymised grades.xlsx"
)
DEFAULT_SAMPLE_PATHS = [
    Path(
        r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040\Sadik Ulusow_1494125_assignsubmission_file\Economics and Society Coursework (60^L1) (Final).pdf"
    ),
    Path(
        r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040\Federica Alvarez-Ossorio Zaldo_1494127_assignsubmission_file\Econ and society FINAL.pdf"
    ),
    Path(
        r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040\Aya Kalada_1493952_assignsubmission_file\EC3040- EconSoc CW2 Final.docx"
    ),
    Path(
        r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040\Apisan Rasiah_1494186_assignsubmission_file\Economics and Society individual coursework (2).docx"
    ),
]
DEFAULT_OUTPUT_DIR = ROOT_DIR / "output" / "human_benchmark_ec3040"
MODEL_NAME = "qwen2:7b"


@dataclass
class HumanRecord:
    participant_id: str
    full_name: str
    total_mark: float
    feedback_comments: str
    part_breakdown: dict[str, float]


@dataclass
class BenchmarkResult:
    variant: str
    run_index: int
    path: str
    participant_id: str
    full_name: str
    file_type: str
    human_total: float
    model_total: float
    total_delta: float
    abs_total_delta: float
    human_part_breakdown: dict[str, float]
    model_part_breakdown: dict[str, float]
    part_abs_delta_sum: float
    part_abs_delta_mean: float
    detected_part_labels: list[str]
    scoring_word_count: int
    structure_word_count: int
    preparation_seconds: float
    extraction_seconds: float
    latency_total: float
    latency_wall_total: float
    latency_structure_detect: float
    latency_part_refine: float
    latency_part_analysis: float
    latency_moderation: float
    latency_finalize: float
    validation_notes: list[str]
    human_feedback_excerpt: str
    model_feedback_excerpt: str


@dataclass
class DebugStageRecord:
    stage: str
    seconds: float
    detail: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run a reproducible EC3040 benchmark against human marks.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--sample", action="append", default=[], help="Submission path. May be passed multiple times.")
    parser.add_argument(
        "--variant",
        action="append",
        choices=("cached_map_only", "prepared_artifact_enriched"),
        default=[],
        help="Benchmark variant to run. May be passed multiple times. Defaults to both.",
    )
    parser.add_argument("--repeats", type=int, default=1, help="Number of repeated runs per script and variant.")
    parser.add_argument(
        "--verifier-model",
        default=None,
        help="Optional verifier model name used during assessment preparation, for example qwen2:7b.",
    )
    parser.add_argument("--debug-script", action="store_true", help="Run one script with detailed stage timing logs.")
    return parser.parse_args()


def load_context() -> MarkingContext:
    if LAST_INGEST_MANIFEST.exists():
        manifest = json.loads(LAST_INGEST_MANIFEST.read_text(encoding="utf-8"))
        documents = manifest.get("documents", {})
        return MarkingContext(
            rubric_text=read_paths_text(tuple(Path(path) for path in documents.get("rubric_files", []))),
            brief_text=read_paths_text(tuple(Path(path) for path in documents.get("brief_files", []))),
            marking_scheme_text=read_paths_text(tuple(Path(path) for path in documents.get("marking_scheme_files", []))),
            graded_sample_text=read_paths_text(tuple(Path(path) for path in documents.get("graded_sample_files", []))),
            other_context_text=read_paths_text(tuple(Path(path) for path in documents.get("other_files", []))),
            max_mark=float(manifest.get("manual_max_mark", 100.0)),
        )
    raise FileNotFoundError(f"Benchmark context manifest not found: {LAST_INGEST_MANIFEST}")


def parse_part_breakdown(feedback_comments: str) -> dict[str, float]:
    pattern = re.compile(r"(Part\s+\d+(?:\s+Q\d+)?):\s*([0-9]+(?:\.[0-9]+)?)/[0-9]+(?:\.[0-9]+)?", re.IGNORECASE)
    breakdown: dict[str, float] = {}
    for label, score in pattern.findall(feedback_comments or ""):
        cleaned_label = re.sub(r"\s+", " ", label.strip()).title().replace("Q", "Q")
        breakdown[cleaned_label] = float(score)
    return breakdown


def load_human_records(workbook_path: Path) -> dict[str, HumanRecord]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Regular submission point"]
    records: dict[str, HumanRecord] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        identifier, full_name, _status, grade, feedback_comments = row[:5]
        if not identifier or grade is None:
            continue
        participant_id_match = re.search(r"(\d{7})", str(identifier))
        if not participant_id_match:
            continue
        participant_id = participant_id_match.group(1)
        records[participant_id] = HumanRecord(
            participant_id=participant_id,
            full_name=str(full_name or "").strip(),
            total_mark=float(grade),
            feedback_comments=str(feedback_comments or "").strip(),
            part_breakdown=parse_part_breakdown(str(feedback_comments or "")),
        )
    return records


def participant_id_from_path(path: Path) -> str:
    match = re.search(r"_(\d{7})_", str(path))
    if not match:
        raise ValueError(f"Could not infer participant id from path: {path}")
    return match.group(1)


def model_part_breakdown(parts: list[Any], part_analyses: list[dict[str, Any]]) -> dict[str, float]:
    resolved: dict[str, float] = {}
    for part, analysis in zip(parts, part_analyses):
        score = analysis.get("provisional_score")
        if score is None and analysis.get("provisional_score_0_to_100") is not None and part.max_mark is not None:
            score = (float(analysis["provisional_score_0_to_100"]) * float(part.max_mark)) / 100.0
        if score is not None:
            resolved[str(part.label)] = round(float(score), 2)
    return resolved


def compare_part_breakdowns(human: dict[str, float], model: dict[str, float]) -> tuple[float, float]:
    labels = sorted(set(human) | set(model))
    if not labels:
        return 0.0, 0.0
    abs_deltas = [abs(float(model.get(label, 0.0)) - float(human.get(label, 0.0))) for label in labels]
    return round(sum(abs_deltas), 2), round(sum(abs_deltas) / len(abs_deltas), 2)


def feedback_excerpt(text: str, limit: int = 220) -> str:
    stripped = re.sub(r"\s+", " ", text.strip())
    return stripped[:limit] + ("..." if len(stripped) > limit else "")


def resolve_variants(args: argparse.Namespace) -> list[str]:
    return args.variant or ["cached_map_only", "prepared_artifact_enriched"]


def apply_variant_to_parts(
    parts: list[Any],
    prepared_assessment_map: Any,
    variant: str,
) -> list[Any]:
    enriched_parts = apply_assessment_map_to_submission_parts(parts, prepared_assessment_map.prepared_map)
    if variant == "prepared_artifact_enriched":
        return apply_prepared_artifact_to_submission_parts(enriched_parts, prepared_assessment_map)
    return enriched_parts


def run_single_submission(
    path: Path,
    context: MarkingContext,
    prepared_assessment_map: Any,
    preparation_seconds: float,
    human_record: HumanRecord,
    variant: str,
    run_index: int,
) -> BenchmarkResult:
    extraction_started = time.perf_counter()
    script_text, structure_text = build_submission_texts_from_path(path)
    extraction_seconds = round(time.perf_counter() - extraction_started, 2)
    scoring_started = time.perf_counter()
    structure_input_text = structure_text or script_text

    stage_started = time.perf_counter()
    detected_parts = detect_submission_parts(
        structure_input_text,
        path.name,
        MODEL_NAME,
        context=context,
        ollama_url=DEFAULT_OLLAMA_URL,
    )
    latency_structure_detect = round(time.perf_counter() - stage_started, 2)

    stage_started = time.perf_counter()
    parts = segment_submission_parts(script_text, detected_parts)
    parts = refine_submission_granularity(parts, context, path.name, MODEL_NAME, DEFAULT_OLLAMA_URL)
    parts = apply_variant_to_parts(parts, prepared_assessment_map, variant)
    latency_part_refine = round(time.perf_counter() - stage_started, 2)

    diagnostics = build_submission_diagnostics(script_text, parts)

    stage_started = time.perf_counter()
    part_analyses: list[dict[str, Any]] = []
    for part in parts:
        if not part.section_text.strip():
            part_analyses.append(build_missing_part_analysis(part))
            continue
        part_analyses.append(
            _run_part_analysis_with_retry(
                part=part,
                context=context,
                filename=path.name,
                model_name=MODEL_NAME,
                ollama_url=DEFAULT_OLLAMA_URL,
            )
        )
    latency_part_analysis = round(time.perf_counter() - stage_started, 2)

    stage_started = time.perf_counter()
    if len(part_analyses) > 1:
        part_analyses = moderate_linked_part_analyses(
            script_text=script_text,
            context=context,
            filename=path.name,
            parts=parts,
            part_analyses=part_analyses,
            assessment_map=prepared_assessment_map.prepared_map,
            model_name=MODEL_NAME,
            ollama_url=DEFAULT_OLLAMA_URL,
        )
    latency_moderation = round(time.perf_counter() - stage_started, 2)

    stage_started = time.perf_counter()
    result = build_local_final_result(
        context=context,
        diagnostics=diagnostics,
        parts=parts,
        part_analyses=part_analyses,
    )
    latency_finalize = round(time.perf_counter() - stage_started, 2)
    latency_total = round(time.perf_counter() - scoring_started, 2)
    latency_wall_total = round(preparation_seconds + extraction_seconds + latency_total, 2)

    model_total = float(result["total_mark"])
    human_total = float(human_record.total_mark)
    model_breakdown = model_part_breakdown(parts, part_analyses)
    part_abs_delta_sum, part_abs_delta_mean = compare_part_breakdowns(human_record.part_breakdown, model_breakdown)
    return BenchmarkResult(
        variant=variant,
        run_index=run_index,
        path=str(path),
        participant_id=human_record.participant_id,
        full_name=human_record.full_name,
        file_type=path.suffix.lower(),
        human_total=human_total,
        model_total=model_total,
        total_delta=round(model_total - human_total, 2),
        abs_total_delta=round(abs(model_total - human_total), 2),
        human_part_breakdown=human_record.part_breakdown,
        model_part_breakdown=model_breakdown,
        part_abs_delta_sum=part_abs_delta_sum,
        part_abs_delta_mean=part_abs_delta_mean,
        detected_part_labels=[part.label for part in parts],
        scoring_word_count=len(script_text.split()),
        structure_word_count=len(structure_input_text.split()),
        preparation_seconds=preparation_seconds,
        extraction_seconds=extraction_seconds,
        latency_total=latency_total,
        latency_wall_total=latency_wall_total,
        latency_structure_detect=latency_structure_detect,
        latency_part_refine=latency_part_refine,
        latency_part_analysis=latency_part_analysis,
        latency_moderation=latency_moderation,
        latency_finalize=latency_finalize,
        validation_notes=list(result.get("validation_notes", [])),
        human_feedback_excerpt=feedback_excerpt(human_record.feedback_comments),
        model_feedback_excerpt=feedback_excerpt(str(result.get("overall_feedback", ""))),
    )


def log_debug_stage(records: list[DebugStageRecord], stage: str, started: float, detail: str = "") -> float:
    seconds = round(time.perf_counter() - started, 2)
    record = DebugStageRecord(stage=stage, seconds=seconds, detail=detail)
    records.append(record)
    suffix = f" :: {detail}" if detail else ""
    print(f"[debug] {stage}: {seconds:.2f}s{suffix}", flush=True)
    return seconds


def run_single_submission_debug(
    path: Path,
    context: MarkingContext,
    prepared_assessment_map: Any,
    preparation_seconds: float,
    human_record: HumanRecord,
    variant: str,
) -> dict[str, Any]:
    debug_records: list[DebugStageRecord] = []

    started = time.perf_counter()
    script_text, structure_text = build_submission_texts_from_path(path)
    extraction_seconds = log_debug_stage(
        debug_records,
        "extract_submission_text",
        started,
        f"scoring_words={len(script_text.split())}, structure_words={len((structure_text or script_text).split())}",
    )
    structure_input_text = structure_text or script_text

    scoring_started = time.perf_counter()
    structure_detection_mode = describe_structure_detection_mode(structure_input_text, context)

    started = time.perf_counter()
    detected_parts = detect_submission_parts(
        structure_input_text,
        path.name,
        MODEL_NAME,
        context=context,
        ollama_url=DEFAULT_OLLAMA_URL,
    )
    structure_detect_seconds = log_debug_stage(
        debug_records,
        "detect_submission_parts",
        started,
        f"mode={structure_detection_mode['mode']}, detected_parts={len(detected_parts)}",
    )

    started = time.perf_counter()
    parts = segment_submission_parts(script_text, detected_parts)
    segment_seconds = log_debug_stage(
        debug_records,
        "segment_submission_parts",
        started,
        f"segmented_parts={len(parts)}",
    )

    started = time.perf_counter()
    parts = refine_submission_granularity(parts, context, path.name, MODEL_NAME, DEFAULT_OLLAMA_URL)
    refine_seconds = log_debug_stage(
        debug_records,
        "refine_submission_granularity",
        started,
        f"refined_parts={len(parts)}",
    )

    started = time.perf_counter()
    parts = apply_variant_to_parts(parts, prepared_assessment_map, variant)
    apply_seconds = log_debug_stage(
        debug_records,
        "apply_assessment_artifact",
        started,
        f"variant={variant}",
    )

    started = time.perf_counter()
    diagnostics = build_submission_diagnostics(script_text, parts)
    diagnostics_seconds = log_debug_stage(
        debug_records,
        "build_submission_diagnostics",
        started,
        f"detected_part_count={diagnostics.detected_part_count}",
    )

    part_analyses: list[dict[str, Any]] = []
    part_debug_rows: list[dict[str, Any]] = []
    for index, part in enumerate(parts, start=1):
        started = time.perf_counter()
        if not part.section_text.strip():
            analysis = build_missing_part_analysis(part)
            seconds = log_debug_stage(
                debug_records,
                "part_analysis_missing",
                started,
                f"{index}/{len(parts)} {part.label}",
            )
        else:
            analysis = _run_part_analysis_with_retry(
                part=part,
                context=context,
                filename=path.name,
                model_name=MODEL_NAME,
                ollama_url=DEFAULT_OLLAMA_URL,
            )
            seconds = log_debug_stage(
                debug_records,
                "part_analysis_model",
                started,
                f"{index}/{len(parts)} {part.label}",
            )
        part_analyses.append(analysis)
        part_debug_rows.append(
            {
                "index": index,
                "label": part.label,
                "seconds": seconds,
                "word_count": len(part.section_text.split()),
                "score": analysis.get("provisional_score"),
                "score_pct": analysis.get("provisional_score_0_to_100"),
            }
        )

    moderation_debug_rows: list[dict[str, Any]] = []
    started = time.perf_counter()
    if len(part_analyses) > 1:
        moderation_plan = describe_moderation_plan(parts, part_analyses, prepared_assessment_map.prepared_map)
        grouped_started = time.perf_counter()
        moderated = moderate_linked_part_analyses(
            script_text=script_text,
            context=context,
            filename=path.name,
            parts=parts,
            part_analyses=part_analyses,
            assessment_map=prepared_assessment_map.prepared_map,
            model_name=MODEL_NAME,
            ollama_url=DEFAULT_OLLAMA_URL,
        )
        moderation_seconds = log_debug_stage(
            debug_records,
            "moderate_linked_part_analyses",
            grouped_started,
            "; ".join(
                f"{item['dependency_group']}={item['reason']}"
                for item in moderation_plan
            ) or "no_linked_groups",
        )
        for item in moderation_plan:
            moderation_debug_rows.append(
                {
                    "seconds": moderation_seconds,
                    "part_count": len(parts),
                    "dependency_group": item["dependency_group"],
                    "part_labels": item["part_labels"],
                    "should_moderate": item["should_moderate"],
                    "reason": item["reason"],
                }
            )
        part_analyses = moderated
    else:
        moderation_seconds = log_debug_stage(debug_records, "moderate_linked_part_analyses", started, "skipped")

    started = time.perf_counter()
    result = build_local_final_result(
        context=context,
        diagnostics=diagnostics,
        parts=parts,
        part_analyses=part_analyses,
    )
    finalize_seconds = log_debug_stage(debug_records, "build_local_final_result", started)

    scoring_total_seconds = round(time.perf_counter() - scoring_started, 2)
    total_wall_seconds = round(extraction_seconds + preparation_seconds + scoring_total_seconds, 2)
    model_breakdown = model_part_breakdown(parts, part_analyses)
    part_abs_delta_sum, part_abs_delta_mean = compare_part_breakdowns(human_record.part_breakdown, model_breakdown)
    payload = {
        "path": str(path),
        "participant_id": human_record.participant_id,
        "full_name": human_record.full_name,
        "variant": variant,
        "preparation_seconds": preparation_seconds,
        "extraction_seconds": extraction_seconds,
        "structure_detect_seconds": structure_detect_seconds,
        "structure_detection_mode": structure_detection_mode,
        "segment_seconds": segment_seconds,
        "refine_seconds": refine_seconds,
        "apply_artifact_seconds": apply_seconds,
        "diagnostics_seconds": diagnostics_seconds,
        "moderation_seconds": moderation_seconds,
        "finalize_seconds": finalize_seconds,
        "scoring_total_seconds": scoring_total_seconds,
        "total_wall_seconds": total_wall_seconds,
        "detected_part_labels": [part.label for part in parts],
        "human_total": human_record.total_mark,
        "model_total": result["total_mark"],
        "abs_total_delta": round(abs(float(result["total_mark"]) - float(human_record.total_mark)), 2),
        "part_abs_delta_sum": part_abs_delta_sum,
        "part_abs_delta_mean": part_abs_delta_mean,
        "validation_notes": list(result.get("validation_notes", [])),
        "part_debug_rows": part_debug_rows,
        "moderation_debug_rows": moderation_debug_rows,
        "debug_stages": [asdict(record) for record in debug_records],
    }
    return payload


def build_aggregate(results: list[BenchmarkResult]) -> dict[str, Any]:
    if not results:
        return {}
    return {
        "sample_count": len(results),
        "mean_abs_total_delta": round(sum(item.abs_total_delta for item in results) / len(results), 2),
        "mean_abs_part_delta": round(sum(item.part_abs_delta_mean for item in results) / len(results), 2),
        "mean_latency_total": round(sum(item.latency_total for item in results) / len(results), 2),
        "mean_latency_wall_total": round(sum(item.latency_wall_total for item in results) / len(results), 2),
        "by_type": {
            file_type: {
                "count": sum(1 for item in results if item.file_type == file_type),
                "mean_abs_total_delta": round(
                    sum(item.abs_total_delta for item in results if item.file_type == file_type)
                    / max(sum(1 for item in results if item.file_type == file_type), 1),
                    2,
                ),
            }
            for file_type in sorted({item.file_type for item in results})
        },
    }


def build_summary_rows(results: list[BenchmarkResult]) -> list[dict[str, Any]]:
    if not results:
        return []
    rows: list[dict[str, Any]] = []
    for variant in sorted({item.variant for item in results}):
        variant_results = [item for item in results if item.variant == variant]
        sample_count = len(variant_results)
        mean_preparation = sum(item.preparation_seconds for item in variant_results) / sample_count
        mean_extraction = sum(item.extraction_seconds for item in variant_results) / sample_count
        mean_script_latency = sum(item.latency_total for item in variant_results) / sample_count
        for batch_size in (1, 5, 10, 25, 50):
            rows.append(
                {
                    "variant": variant,
                    "sample_count": sample_count,
                    "repeat_count": len({item.run_index for item in variant_results}),
                    "mean_abs_total_delta": round(sum(item.abs_total_delta for item in variant_results) / sample_count, 2),
                    "mean_abs_part_delta": round(sum(item.part_abs_delta_mean for item in variant_results) / sample_count, 2),
                    "mean_model_total": round(sum(item.model_total for item in variant_results) / sample_count, 2),
                    "mean_human_total": round(sum(item.human_total for item in variant_results) / sample_count, 2),
                    "mean_extraction_seconds": round(mean_extraction, 2),
                    "mean_script_latency_seconds": round(mean_script_latency, 2),
                    "mean_preparation_seconds": round(mean_preparation, 2),
                    "amortized_batch_size": batch_size,
                    "amortized_total_seconds_per_script": round(mean_extraction + mean_script_latency + mean_preparation / max(batch_size, 1), 2),
                }
            )
    return rows


def build_variant_comparison_rows(results: list[BenchmarkResult]) -> list[dict[str, Any]]:
    variants = sorted({item.variant for item in results})
    if len(variants) < 2:
        return []
    keyed = {
        (item.variant, item.run_index, item.participant_id, item.path): item
        for item in results
    }
    baseline = "cached_map_only"
    candidate = "prepared_artifact_enriched"
    rows: list[dict[str, Any]] = []
    for run_index in sorted({item.run_index for item in results}):
        for item in [entry for entry in results if entry.variant == baseline and entry.run_index == run_index]:
            other = keyed.get((candidate, run_index, item.participant_id, item.path))
            if other is None:
                continue
            rows.append(
                {
                    "run_index": run_index,
                    "participant_id": item.participant_id,
                    "path": item.path,
                    "baseline_variant": baseline,
                    "candidate_variant": candidate,
                    "abs_total_delta_improvement": round(item.abs_total_delta - other.abs_total_delta, 2),
                    "abs_part_delta_improvement": round(item.part_abs_delta_mean - other.part_abs_delta_mean, 2),
                    "script_latency_delta_seconds": round(other.latency_total - item.latency_total, 2),
                    "preparation_latency_delta_seconds": round(other.preparation_seconds - item.preparation_seconds, 2),
                }
            )
    return rows


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sample_paths = [Path(item) for item in args.sample] if args.sample else list(DEFAULT_SAMPLE_PATHS)
    variants = resolve_variants(args)
    human_records = load_human_records(workbook_path)
    context = load_context()

    if args.debug_script:
        if not sample_paths:
            raise SystemExit("Provide at least one sample path for --debug-script.")
        debug_path = sample_paths[0]
        debug_variant = variants[0]
        clear_prepared_assessment_map_cache()
        prep_started = time.perf_counter()
        prepared_assessment_map = prepare_assessment_map(
            context=context,
            verifier_model_name=args.verifier_model,
        )
        preparation_seconds = round(time.perf_counter() - prep_started, 2)
        participant_id = participant_id_from_path(debug_path)
        human_record = human_records[participant_id]
        print(f"[debug] script={debug_path}")
        print(f"[debug] variant={debug_variant}")
        print(f"[debug] preparation_seconds={preparation_seconds:.2f}s", flush=True)
        payload = run_single_submission_debug(
            path=debug_path,
            context=context,
            prepared_assessment_map=prepared_assessment_map,
            preparation_seconds=preparation_seconds,
            human_record=human_record,
            variant=debug_variant,
        )
        (output_dir / "debug_run.json").write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
        pd.DataFrame(payload["part_debug_rows"]).to_csv(output_dir / "debug_parts.csv", index=False)
        pd.DataFrame(payload["debug_stages"]).to_csv(output_dir / "debug_stages.csv", index=False)
        print(json.dumps(payload, ensure_ascii=True, indent=2))
        return

    results: list[BenchmarkResult] = []
    preparation_by_variant: dict[str, float] = {}
    for variant in variants:
        clear_prepared_assessment_map_cache()
        prep_started = time.perf_counter()
        prepared_assessment_map = prepare_assessment_map(
            context=context,
            verifier_model_name=args.verifier_model,
        )
        preparation_seconds = round(time.perf_counter() - prep_started, 2)
        preparation_by_variant[variant] = preparation_seconds
        for run_index in range(1, args.repeats + 1):
            for index, path in enumerate(sample_paths, start=1):
                participant_id = participant_id_from_path(path)
                human_record = human_records[participant_id]
                print(f"[{variant} run {run_index} {index}/{len(sample_paths)}] {path.name} :: {human_record.full_name}")
                row = run_single_submission(
                    path=path,
                    context=context,
                    prepared_assessment_map=prepared_assessment_map,
                    preparation_seconds=preparation_seconds,
                    human_record=human_record,
                    variant=variant,
                    run_index=run_index,
                )
                results.append(row)
                print(json.dumps(asdict(row), ensure_ascii=True))

    results_df = pd.DataFrame(asdict(item) for item in results)
    summary_rows = build_summary_rows(results)
    comparison_rows = build_variant_comparison_rows(results)
    payload = {
        "assessment_context": str(ASSESSMENT_CONTEXT_PDF),
        "workbook": str(workbook_path),
        "model_name": MODEL_NAME,
        "verifier_model_name": args.verifier_model,
        "variants": variants,
        "repeats": args.repeats,
        "prepared_assessment_seconds_by_variant": preparation_by_variant,
        "sample_paths": [str(path) for path in sample_paths],
        "aggregate": {
            variant: build_aggregate([item for item in results if item.variant == variant])
            for variant in variants
        },
        "summary_rows": summary_rows,
        "comparison_rows": comparison_rows,
        "results": [asdict(item) for item in results],
    }
    (output_dir / "summary.json").write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
    (output_dir / "sample_paths.txt").write_text("\n".join(str(path) for path in sample_paths), encoding="utf-8")
    results_df.to_csv(output_dir / "runs.csv", index=False)
    pd.DataFrame(summary_rows).to_csv(output_dir / "summary.csv", index=False)
    if comparison_rows:
        pd.DataFrame(comparison_rows).to_csv(output_dir / "variant_comparison.csv", index=False)


if __name__ == "__main__":
    main()
