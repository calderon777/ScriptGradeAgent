from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
WORKBOOK_PATH = Path(r"C:\Users\Cam\OneDrive - City St George's, University of London\grading test\test 2\EC3040 deanonymised grades.xlsx")
EXEMPLAR_NAMES = [
    "Sadik Ulusow",
    "Federica Alvarez-Ossorio Zaldo",
    "Alessandro Ungurean",
    "Fahima Ali",
]
PART_SCORE_RE = re.compile(r"(Part\s+\d(?:\s+Q\d)?):\s*([0-9]+(?:\.[0-9]+)?)/([0-9]+(?:\.[0-9]+)?)")


@dataclass
class WorkbookRow:
    full_name: str
    grade: float | None
    feedback: str
    part_scores: dict[str, dict[str, float]]


def parse_part_scores(text: str) -> dict[str, dict[str, float]]:
    scores: dict[str, dict[str, float]] = {}
    for part, score, max_score in PART_SCORE_RE.findall(text or ""):
        scores[part] = {"score": float(score), "max_score": float(max_score)}
    return scores


def extract_exemplar_rows() -> list[WorkbookRow]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    rows: list[WorkbookRow] = []
    for ws in wb.worksheets:
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        h = {str(v): i for i, v in enumerate(headers)}
        for values in ws.iter_rows(min_row=2, values_only=True):
            full_name = str(values[h["Full name"]] or "")
            if not any(name in full_name for name in EXEMPLAR_NAMES):
                continue
            feedback = str(values[h["Feedback comments"]] or "").strip()
            grade = values[h["Grade"]]
            rows.append(
                WorkbookRow(
                    full_name=full_name,
                    grade=float(grade) if isinstance(grade, (int, float)) else None,
                    feedback=feedback,
                    part_scores=parse_part_scores(feedback),
                )
            )
    return rows


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages).strip()


def extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml_bytes = zf.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    return "\n".join(paragraphs).strip()


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    exemplars = [asdict(row) for row in extract_exemplar_rows()]
    (OUTPUT_DIR / "human_exemplars.json").write_text(json.dumps(exemplars, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "ec3040_marking_scheme.txt").write_text(
        extract_pdf_text(INPUT_DIR / "ec3040_marking_scheme.pdf"),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "maria_submission.txt").write_text(
        extract_pdf_text(INPUT_DIR / "maria_submission.pdf"),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "alice_submission.txt").write_text(
        extract_docx_text(INPUT_DIR / "alice_submission.docx"),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
