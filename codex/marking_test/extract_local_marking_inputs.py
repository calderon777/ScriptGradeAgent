from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
WORKBOOK_PATH = INPUT_DIR / "EC3040_deanonymised_grades.xlsx"
TARGETS = {
    "sadik": INPUT_DIR / "sadik_submission.pdf",
    "federica": INPUT_DIR / "federica_submission.pdf",
}
PART_SCORE_RE = re.compile(r"(Part\s+\d(?:\s+Q\d)?):\s*([0-9]+(?:\.[0-9]+)?)/([0-9]+(?:\.[0-9]+)?)")


@dataclass
class WorkbookRow:
    sheet: str
    identifier: str
    full_name: str
    grade: float | None
    feedback: str
    part_scores: dict[str, dict[str, float]]


def parse_part_scores(text: str) -> dict[str, dict[str, float]]:
    scores: dict[str, dict[str, float]] = {}
    for part, score, max_score in PART_SCORE_RE.findall(text or ""):
        scores[part] = {"score": float(score), "max_score": float(max_score)}
    return scores


def extract_workbook_rows() -> list[WorkbookRow]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    rows: list[WorkbookRow] = []
    for ws in wb.worksheets:
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {str(value): index for index, value in enumerate(headers)}
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(values):
                continue
            full_name = str(values[header_map["Full name"]] or "").strip()
            feedback = str(values[header_map["Feedback comments"]] or "").strip()
            grade = values[header_map["Grade"]]
            rows.append(
                WorkbookRow(
                    sheet=ws.title,
                    identifier=str(values[header_map["Identifier"]] or "").strip(),
                    full_name=full_name,
                    grade=float(grade) if isinstance(grade, (int, float)) else None,
                    feedback=feedback,
                    part_scores=parse_part_scores(feedback),
                )
            )
    return rows


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages: list[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n\n".join(pages).strip()


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook_rows = extract_workbook_rows()
    target_rows = [
        asdict(row)
        for row in workbook_rows
        if "Sadik Ulusow" in row.full_name or "Federica Alvarez-Ossorio Zaldo" in row.full_name
    ]
    non_target_rows = [
        asdict(row)
        for row in workbook_rows
        if "Sadik Ulusow" not in row.full_name and "Federica Alvarez-Ossorio Zaldo" not in row.full_name
    ]

    (OUTPUT_DIR / "target_rows.json").write_text(json.dumps(target_rows, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "non_target_rows.json").write_text(json.dumps(non_target_rows, indent=2), encoding="utf-8")

    for name, path in TARGETS.items():
        text = extract_pdf_text(path)
        (OUTPUT_DIR / f"{name}_submission.txt").write_text(text, encoding="utf-8")


if __name__ == "__main__":
    main()
